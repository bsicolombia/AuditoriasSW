from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import get_template
import pandas as pd
from xhtml2pdf import pisa
from .models import Auditoria
from django.db.models import Count
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

def cargar_excel(request):

    if request.method == "POST":

        archivo = request.FILES["archivo"]

        excel = pd.read_excel(archivo)

        for _, fila in excel.iterrows():

            Auditoria.objects.create(
                fecha=fila["Fecha"],
                nombre_auditor=fila["Auditor"],
                numero_cedula=fila["Cedula"],
                aplicativo=fila["Aplicativo"],
                fecha_operacion=fila["Fecha Operacion"],
                nombre_tecnico=fila["Tecnico"],
                numero_cuenta_contrato=fila["Cuenta"],
                numero_orden=fila["Orden"],
                tipo_operacion=fila["Tipo Operacion"],
                resultado_auditoria=fila["Resultado"],
                observacion=fila["Observacion"],
                tipo_hallazgo=fila["Tipo Hallazgo"],
                hallazgo=fila["Hallazgo"]
            )

        return render(
            request,
            "carga/carga.html",
            {
                "mensaje": "Excel cargado correctamente"
            }
        )

    return render(
        request,
        "carga/carga.html"
    )


def filtrar_auditorias(request):

    auditorias = Auditoria.objects.all()

    # ==========================================
    # FILTROS NORMALES
    # ==========================================

    filtros = {

        "fecha": "fecha",

        "nombre_auditor": "nombre_auditor__icontains",

        "numero_cedula": "numero_cedula__icontains",

        "aplicativo": "aplicativo__icontains",

        "fecha_operacion": "fecha_operacion",

        "nombre_tecnico": "nombre_tecnico__icontains",

        "numero_cuenta_contrato": "numero_cuenta_contrato__icontains",

        "numero_orden": "numero_orden__icontains",
        
        "tipo_operacion": "tipo_operacion",

        "resultado_auditoria": "resultado_auditoria",

        "tipo_hallazgo": "tipo_hallazgo",

        "hallazgo": "hallazgo__icontains",
    }

    # ==========================================
    # APLICAR FILTROS NORMALES
    # ==========================================

    for parametro, campo in filtros.items():

        valor = request.GET.get(parametro)

        if valor:

            auditorias = auditorias.filter(
                **{campo: valor}
            )

    # ==========================================
    # FECHA OPERACIÓN DESDE
    # ==========================================

    fecha_inicio = request.GET.get("fecha_inicio")

    if fecha_inicio:

        auditorias = auditorias.filter(
            fecha__gte=fecha_inicio
        )

    # ==========================================
    # FECHA OPERACIÓN HASTA
    # ==========================================

    fecha_fin = request.GET.get("fecha_fin")

    if fecha_fin:

        auditorias = auditorias.filter(
            fecha__lte=fecha_fin
        )

    # ==========================================
    # FECHA AUDITORÍA DESDE
    # Campo: fecha
    # ==========================================

    fecha_inicio_auditoria = request.GET.get(
        "fecha_inicio_auditoria"
    )

    if fecha_inicio_auditoria:

        auditorias = auditorias.filter(
            fecha__gte=fecha_inicio_auditoria
        )

    # ==========================================
    # FECHA AUDITORÍA HASTA
    # Campo: fecha
    # ==========================================

    fecha_fin_auditoria = request.GET.get(
        "fecha_fin_auditoria"
    )

    if fecha_fin_auditoria:

        auditorias = auditorias.filter(
            fecha__lte=fecha_fin_auditoria
        )
    
    return auditorias


def consulta(request):

    auditorias = filtrar_auditorias(request)
    cantidad = auditorias.count()

    return render(
        request,
        "auditorias/consulta.html",

        {
            "Auditorias": auditorias,
            "Cantidad": cantidad
        }
    )


def generate_pdf(request):

    auditorias = filtrar_auditorias(request)

    cantidad = auditorias.count()

    # Evitar generar PDFs demasiado grandes
    if cantidad > 50000:

        return HttpResponse(
            f"""
            <html>
                <head>
                    <meta charset="UTF-8">
                    <title>Demasiados registros</title>
                </head>

                <body>

                    <h2>Demasiados registros para generar el PDF</h2>

                    <p>
                        La consulta contiene
                        <strong>{cantidad}</strong>
                        registros.
                    </p>

                    <p>
                        Por favor, aplique uno o varios filtros
                        antes de generar el PDF.
                    </p>

                    <a href="/auditorias/consulta/">
                        Volver a consultas
                    </a>

                </body>
            </html>
            """,
            status=400
        )

    template = get_template(
        "auditorias/pdf_auditorias.html"
    )

    context = {
        "Auditorias": auditorias
    }

    html = template.render(context)

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="auditorias.pdf"'
    )

    pisa_status = pisa.CreatePDF(
        html,
        dest=response
    )

    if pisa_status.err:

        return HttpResponse(
            "Error al generar el PDF",
            status=500
        )

    return response

def exportar_excel(request):
    auditorias = filtrar_auditorias(request)
    datos =[]
    
    for a in auditorias:
        datos.append({
            "Fecha": a.fecha,
            "Nombre Auditor": a.nombre_auditor,
            "Numero Cedula": a.numero_cedula,
            "Aplicativo": a.aplicativo,
            "Fecha Operacion": a.fecha_operacion,
            "Nombre Tecnico": a.nombre_tecnico,
            "Numero Cuenta Contrato": a.numero_cuenta_contrato,
            "Numero Orden": a.numero_orden,
            "Tipo Operacion": a.tipo_operacion,
            "Resultado Auditoria": a.resultado_auditoria,
            "Observacion": a.observacion,
            "Tipo Hallazgo": a.tipo_hallazgo,
            "Hallazgo": a.hallazgo
        })
    
    df = pd.DataFrame(datos)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    response["Content-Disposition"] = (
        'attachment; filename="auditorias.xlsx"'
    )
    
    with pd.ExcelWriter(
        response,
        engine="openpyxl"
    ) as writer:
        
        df.to_excel(
            writer,
            index=False,
            sheet_name="Auditorias"
        )
        
        return response
    
def estadistica(request):

    # ==========================================
    # ESTADÍSTICAS DE OPERACIONES
    # ==========================================

    suspensiones = Auditoria.objects.filter(
        tipo_operacion="DC00"
    ).count()

    reconexiones = Auditoria.objects.filter(
        tipo_operacion="RC00"
    ).count()

    zvcl = Auditoria.objects.filter(
        tipo_operacion="ZVCL"
    ).count()

    datos_operaciones = {
        "Suspensiones": suspensiones,
        "Reconexiones": reconexiones,
        "Zvcl": zvcl,
    }

    return render(
        request,
        "auditorias/estadistica.html",
        {
            "datos_operaciones": datos_operaciones,
        }
    )



def exportar_estadisticas_excel(request):

    # ==========================================
    # ESTADÍSTICAS GENERALES
    # ==========================================

    cantidad = Auditoria.objects.count()

    suspensiones = Auditoria.objects.filter(
        tipo_operacion="DC00"
    ).count()

    reconexiones = Auditoria.objects.filter(
        tipo_operacion="RC00"
    ).count()

    zvcl = Auditoria.objects.filter(
        tipo_operacion="ZVCL"
    ).count()


    # ==========================================
    # ESTADÍSTICAS POR TÉCNICO
    # ==========================================

    estadisticas_tecnico = list(
        Auditoria.objects
        .values("nombre_tecnico")
        .annotate(total=Count("id"))
        .order_by("-total")
    )


    datos_tecnicos = []

    for tecnico in estadisticas_tecnico:

        datos_tecnicos.append({

            "Nombre Técnico":
                tecnico["nombre_tecnico"],

            "Cantidad de Auditorías":
                tecnico["total"]

        })


    df_tecnicos = pd.DataFrame(
        datos_tecnicos
    )


    # ==========================================
    # ESTADÍSTICAS POR DÍA
    # ==========================================

    estadisticas_dia = list(
        Auditoria.objects
        .values("fecha_operacion")
        .annotate(total=Count("id"))
        .order_by("fecha_operacion")
    )


    datos_dias = []

    for dia in estadisticas_dia:

        datos_dias.append({

            "Fecha":
                dia["fecha_operacion"],

            "Cantidad de Auditorías":
                dia["total"]

        })


    df_dias = pd.DataFrame(
        datos_dias
    )


    # ==========================================
    # RESUMEN
    # ==========================================

    datos_resumen = [

        {
            "Estadística":
                "Total de Auditorías",

            "Cantidad":
                cantidad
        },

        {
            "Estadística":
                "Suspensiones",

            "Cantidad":
                suspensiones
        },

        {
            "Estadística":
                "Reconexiones",

            "Cantidad":
                reconexiones
        },

        {
            "Estadística":
                "ZVCL",

            "Cantidad":
                zvcl
        },

        {
            "Estadística":
                "Total de Técnicos",

            "Cantidad":
                len(estadisticas_tecnico)
        }

    ]


    df_resumen = pd.DataFrame(
        datos_resumen
    )


    # ==========================================
    # CREAR EXCEL
    # ==========================================

    response = HttpResponse(

        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )


    response["Content-Disposition"] = (
        'attachment; filename="estadisticas_auditorias.xlsx"'
    )


    # ==========================================
    # ESCRIBIR LAS 3 HOJAS
    # ==========================================

    with pd.ExcelWriter(
        response,
        engine="openpyxl"
    ) as writer:

        df_resumen.to_excel(

            writer,

            index=False,

            sheet_name="Resumen"

        )


        df_tecnicos.to_excel(

            writer,

            index=False,

            sheet_name="Por Tecnico"

        )


        df_dias.to_excel(

            writer,

            index=False,

            sheet_name="Por Dia"

        )

    return response

from django.http import HttpResponse
from django.db.models import Count
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def exportar_estadisticas(request):

    # ==========================================================
    # QUERYSET GENERAL
    # ==========================================================

    auditorias = Auditoria.objects.all()


    # ==========================================================
    # FILTROS PARA AUDITORÍAS POR DIGITADOR
    # ==========================================================

    anio_digitador = request.GET.get(
        "anio_digitador"
    )

    mes_digitador = request.GET.get(
        "mes_digitador"
    )


    auditorias_digitador = auditorias


    if anio_digitador:

        auditorias_digitador = (
            auditorias_digitador
            .filter(
                fecha_operacion__year=anio_digitador
            )
        )


    if mes_digitador:

        auditorias_digitador = (
            auditorias_digitador
            .filter(
                fecha_operacion__month=mes_digitador
            )
        )


    # ==========================================================
    # FILTROS PARA CANTIDAD DE HALLAZGOS
    # ==========================================================

    anio_hallazgos = request.GET.get(
        "anio_hallazgos"
    )

    mes_hallazgos = request.GET.get(
        "mes_hallazgos"
    )


    auditorias_hallazgos = auditorias


    if anio_hallazgos:

        auditorias_hallazgos = (
            auditorias_hallazgos
            .filter(
                fecha_operacion__year=anio_hallazgos
            )
        )


    if mes_hallazgos:

        auditorias_hallazgos = (
            auditorias_hallazgos
            .filter(
                fecha_operacion__month=mes_hallazgos
            )
        )


    # ==========================================================
    # CREAR LIBRO
    # ==========================================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Estadísticas"


    # ==========================================================
    # COLORES
    # ==========================================================

    COLOR_TITULO = "1F2937"

    COLOR_HEADER = "D9EAF7"

    COLOR_TOTAL = "D9EAD3"

    COLOR_BORDE = "7F8C8D"

    COLOR_TEXTO = "000000"


    # ==========================================================
    # BORDE
    # ==========================================================

    borde = Border(

        left=Side(
            style="thin",
            color=COLOR_BORDE
        ),

        right=Side(
            style="thin",
            color=COLOR_BORDE
        ),

        top=Side(
            style="thin",
            color=COLOR_BORDE
        ),

        bottom=Side(
            style="thin",
            color=COLOR_BORDE
        )

    )


    # ==========================================================
    # FUNCIÓN CREAR TÍTULO
    # ==========================================================

    def crear_titulo(
        ws,
        fila,
        columna_inicio,
        columna_fin,
        titulo,
        color=COLOR_TITULO
    ):

        ws.merge_cells(

            start_row=fila,

            start_column=columna_inicio,

            end_row=fila,

            end_column=columna_fin

        )


        celda = ws.cell(

            fila,

            columna_inicio

        )


        celda.value = titulo


        celda.font = Font(

            bold=True,

            color="FFFFFF",

            size=11

        )


        celda.fill = PatternFill(

            "solid",

            fgColor=color

        )


        celda.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )


        celda.border = borde


        for columna in range(

            columna_inicio,

            columna_fin + 1

        ):

            ws.cell(

                fila,

                columna

            ).fill = PatternFill(

                "solid",

                fgColor=color

            )

            ws.cell(

                fila,

                columna

            ).border = borde


        ws.row_dimensions[fila].height = 22


    # ==========================================================
    # FUNCIÓN CREAR ENCABEZADOS
    # ==========================================================

    def crear_encabezados(
        ws,
        fila,
        columna,
        encabezados
    ):

        for i, encabezado in enumerate(
            encabezados
        ):

            celda = ws.cell(

                fila,

                columna + i

            )


            celda.value = encabezado


            celda.font = Font(

                bold=True,

                color=COLOR_TEXTO

            )


            celda.fill = PatternFill(

                "solid",

                fgColor=COLOR_HEADER

            )


            celda.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )


            celda.border = borde


    # ==========================================================
    # FUNCIÓN ESCRIBIR DATOS
    # ==========================================================

    def escribir_datos(
        ws,
        fila_inicio,
        columna_inicio,
        datos
    ):

        fila = fila_inicio


        for dato in datos:

            for i, valor in enumerate(dato):

                celda = ws.cell(

                    fila,

                    columna_inicio + i

                )


                celda.value = valor


                celda.border = borde


                celda.alignment = Alignment(

                    vertical="center"

                )


                if hasattr(
                    valor,
                    "strftime"
                ):

                    celda.number_format = (
                        "dd/mm/yyyy"
                    )


            fila += 1


        if not datos:

            return fila_inicio - 1


        return fila - 1


    # ==========================================================
    # FUNCIÓN CREAR TOTAL
    # ==========================================================

    def crear_total(
        ws,
        fila,
        columna_inicio,
        valores
    ):

        for i, valor in enumerate(valores):

            celda = ws.cell(

                fila,

                columna_inicio + i

            )


            celda.value = valor


            celda.font = Font(

                bold=True

            )


            celda.fill = PatternFill(

                "solid",

                fgColor=COLOR_TOTAL

            )


            celda.border = borde


            celda.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )


    # ==========================================================
    # FUNCIÓN CREAR TABLA
    # ==========================================================

    def crear_tabla(
        ws,
        nombre,
        fila_inicio,
        columna_inicio,
        fila_fin,
        columna_fin
    ):

        # ------------------------------------------------------
        # NO CREAR TABLA SI NO HAY DATOS
        # ------------------------------------------------------

        if fila_fin < fila_inicio + 1:

            return


        # ------------------------------------------------------
        # REFERENCIA
        # ------------------------------------------------------

        referencia = (

            f"{get_column_letter(columna_inicio)}"

            f"{fila_inicio}:"

            f"{get_column_letter(columna_fin)}"

            f"{fila_fin}"

        )


        # ------------------------------------------------------
        # CREAR TABLA
        # ------------------------------------------------------

        tabla = Table(

            displayName=nombre,

            ref=referencia

        )


        # ------------------------------------------------------
        # ESTILO
        # ------------------------------------------------------

        estilo = TableStyleInfo(

            name="TableStyleMedium2",

            showFirstColumn=False,

            showLastColumn=False,

            showRowStripes=True,

            showColumnStripes=False

        )


        tabla.tableStyleInfo = estilo


        # ------------------------------------------------------
        # AGREGAR
        # ------------------------------------------------------

        ws.add_table(
            tabla
        )


    # ==========================================================
    # FUNCIÓN AJUSTAR ANCHO
    # ==========================================================

    def ajustar_ancho(
        columna_inicio,
        columna_fin,
        ancho_maximo=35
    ):

        for columna in range(

            columna_inicio,

            columna_fin + 1

        ):

            letra = get_column_letter(
                columna
            )


            maximo = 0


            for celda in ws[letra]:

                if celda.value is not None:

                    longitud = len(
                        str(celda.value)
                    )


                    maximo = max(

                        maximo,

                        longitud

                    )


            ws.column_dimensions[
                letra
            ].width = min(

                maximo + 2,

                ancho_maximo

            )


    # ==========================================================
    # CONFIGURACIÓN GENERAL
    # ==========================================================

    ws.sheet_view.showGridLines = True

    ws.freeze_panes = "A3"


    # ==========================================================
    # 1. AUDITORÍAS TOTALES REALIZADAS
    # ==========================================================

    datos = (

        auditorias

        .values(
            "fecha_operacion"
        )

        .annotate(
            total=Count("id")
        )

        .order_by(
            "fecha_operacion"
        )

    )


    datos = list(datos)


    columna = 1


    crear_titulo(

        ws,

        1,

        columna,

        columna + 1,

        "AUDITORÍAS TOTALES REALIZADAS"

    )


    crear_encabezados(

        ws,

        2,

        columna,

        [

            "Fecha",

            "Total"

        ]

    )


    filas = []

    total_auditorias = 0


    for dato in datos:

        filas.append([

            dato["fecha_operacion"],

            dato["total"]

        ])


        total_auditorias += dato["total"]


    ultima_fila_totales = escribir_datos(

        ws,

        3,

        columna,

        filas

    )


    crear_total(

        ws,

        ultima_fila_totales + 1,

        columna,

        [

            "Total general",

            total_auditorias

        ]

    )


    crear_tabla(

        ws,

        "TablaAuditoriasTotales",

        2,

        columna,

        ultima_fila_totales,

        columna + 1

    )


    ajustar_ancho(

        columna,

        columna + 1

    )


    # ==========================================================
    # 2. AUDITORÍAS QUE NO CUMPLEN
    # ==========================================================

    datos = (

        auditorias

        .filter(

            resultado_auditoria="no_cumple"

        )

        .values(

            "fecha_operacion"

        )

        .annotate(

            total=Count("id")

        )

        .order_by(

            "fecha_operacion"

        )

    )


    datos = list(datos)


    columna = 4


    crear_titulo(

        ws,

        1,

        columna,

        columna + 1,

        "AUDITORÍAS QUE NO CUMPLEN"

    )


    crear_encabezados(

        ws,

        2,

        columna,

        [

            "Fecha",

            "Total"

        ]

    )


    filas = []

    total_no_cumplen = 0


    for dato in datos:

        filas.append([

            dato["fecha_operacion"],

            dato["total"]

        ])


        total_no_cumplen += dato["total"]


    ultima_fila_no_cumplen = escribir_datos(

        ws,

        3,

        columna,

        filas

    )


    crear_total(

        ws,

        ultima_fila_no_cumplen + 1,

        columna,

        [

            "Total general",

            total_no_cumplen

        ]

    )


    crear_tabla(

        ws,

        "TablaNoCumplen",

        2,

        columna,

        ultima_fila_no_cumplen,

        columna + 1

    )


    ajustar_ancho(

        columna,

        columna + 1

    )


    # ==========================================================
    # 3. AUDITORÍAS PARA DESCONTAR A LOS TÉCNICOS
    # ==========================================================

    # ----------------------------------------------------------
    # FILTROS PROPIOS DE ESTA SECCIÓN
    # ----------------------------------------------------------

    anio_hallazgos = request.GET.get(
        "anio_hallazgos"
    )

    mes_hallazgos = request.GET.get(
        "mes_hallazgos"
    )


    # ----------------------------------------------------------
    # BASE DE DATOS
    # ----------------------------------------------------------

    auditorias_descontar = Auditoria.objects.all()


    # ----------------------------------------------------------
    # FILTRO POR AÑO
    # ----------------------------------------------------------

    if (
        anio_hallazgos
        and anio_hallazgos != "todos"
    ):

        auditorias_descontar = (
            auditorias_descontar
            .filter(
                fecha_operacion__year=anio_hallazgos
            )
        )


    # ----------------------------------------------------------
    # FILTRO POR MES
    # ----------------------------------------------------------

    if (
        mes_hallazgos
        and mes_hallazgos != "todos"
    ):

        auditorias_descontar = (
            auditorias_descontar
            .filter(
                fecha_operacion__month=mes_hallazgos
            )
        )


    # ----------------------------------------------------------
    # SOLO AUDITORÍAS QUE NO CUMPLEN
    # ----------------------------------------------------------

    auditorias_descontar = (
        auditorias_descontar
        .filter(
            resultado_auditoria="no_cumple"
        )
    )


    # ----------------------------------------------------------
    # OBTENER AUDITORÍAS POR TÉCNICO
    # ----------------------------------------------------------

    datos = list(

        auditorias_descontar
        .exclude(
            nombre_tecnico__isnull=True
        )
        .exclude(
            nombre_tecnico__exact=""
        )
        .values(
            "nombre_tecnico"
        )
        .annotate(
            total=Count("id")
        )
        .order_by(
            "-total",
            "nombre_tecnico"
        )

    )


    # ----------------------------------------------------------
    # COLUMNA
    # ----------------------------------------------------------

    columna = 7


    # ----------------------------------------------------------
    # TÍTULO
    # ----------------------------------------------------------

    crear_titulo(

        ws,

        1,

        columna,

        columna + 1,

        "AUDITORIAS PARA DESCONTAR A LOS TECNICOS"

    )


    # ----------------------------------------------------------
    # INFORMACIÓN DEL MES
    # ----------------------------------------------------------

    celda = ws.cell(
        2,
        columna
    )

    celda.value = "MES"

    celda.font = Font(
        bold=True
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=COLOR_HEADER
    )

    celda.border = borde


    celda = ws.cell(
        2,
        columna + 1
    )


    meses = {

        "1": "ENERO",
        "2": "FEBRERO",
        "3": "MARZO",
        "4": "ABRIL",
        "5": "MAYO",
        "6": "JUNIO",
        "7": "JULIO",
        "8": "AGOSTO",
        "9": "SEPTIEMBRE",
        "10": "OCTUBRE",
        "11": "NOVIEMBRE",
        "12": "DICIEMBRE"

    }


    if (
        mes_hallazgos
        and mes_hallazgos != "todos"
    ):

        celda.value = meses.get(
            str(mes_hallazgos),
            str(mes_hallazgos)
        )

    else:

        celda.value = "(Todos)"


    celda.border = borde


    # ----------------------------------------------------------
    # FECHA
    # ----------------------------------------------------------

    celda = ws.cell(
        3,
        columna
    )

    celda.value = "FECHA"

    celda.font = Font(
        bold=True
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=COLOR_HEADER
    )

    celda.border = borde


    celda = ws.cell(
        3,
        columna + 1
    )

    celda.value = "(Todas)"

    celda.border = borde


    # ----------------------------------------------------------
    # RESULTADO AUDITORÍA
    # ----------------------------------------------------------

    celda = ws.cell(
        4,
        columna
    )

    celda.value = "Resultado Auditoria"

    celda.font = Font(
        bold=True
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=COLOR_HEADER
    )

    celda.border = borde


    celda = ws.cell(
        4,
        columna + 1
    )

    celda.value = "No cumple"

    celda.border = borde


    # ----------------------------------------------------------
    # ENCABEZADOS
    # ----------------------------------------------------------

    fila_encabezado = 6


    crear_encabezados(

        ws,

        fila_encabezado,

        columna,

        [
            "Etiquetas de fila",
            "Cuenta de Resultado Auditoria"
        ]

    )


    # ----------------------------------------------------------
    # ESCRIBIR DATOS
    # ----------------------------------------------------------

    fila_actual = 7


    total_general_descontar = 0


    for dato in datos:

        tecnico = dato[
            "nombre_tecnico"
        ]

        total = dato[
            "total"
        ]


        total_general_descontar += total


        # ------------------------------------------------------
        # NOMBRE DEL TÉCNICO
        # ------------------------------------------------------

        celda = ws.cell(

            fila_actual,

            columna

        )

        celda.value = tecnico

        celda.border = borde

        celda.font = Font(

            bold=False

        )

        celda.alignment = Alignment(

            vertical="center"

        )


        # ------------------------------------------------------
        # CANTIDAD
        # ------------------------------------------------------

        celda = ws.cell(

            fila_actual,

            columna + 1

        )

        celda.value = total

        celda.border = borde

        celda.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )


        fila_actual += 1


    # ----------------------------------------------------------
    # TOTAL GENERAL
    # ----------------------------------------------------------

    crear_total(

        ws,

        fila_actual,

        columna,

        [

            "Total general",

            total_general_descontar

        ]

    )


    # ----------------------------------------------------------
    # GUARDAR ÚLTIMA FILA
    # ----------------------------------------------------------

    ultima_fila_tecnicos = fila_actual


    # ----------------------------------------------------------
    # IMPORTANTE
    #
    # NO USAR crear_tabla() EN ESTA SECCIÓN
    #
    # La tabla queda como celdas normales.
    # Esto evita:
    #
    # Característica quitada: Autofiltro
    # Característica quitada: Tabla
    #
    # ----------------------------------------------------------


    # ----------------------------------------------------------
    # ANCHO
    # ----------------------------------------------------------

    ajustar_ancho(

        columna,

        columna + 1,

        45

    )


    # ==========================================================
    # 4. ERRORES COMETIDOS POR LOS TÉCNICOS
    # ==========================================================

    # ----------------------------------------------------------
    # APLICAR LOS MISMOS FILTROS DE AÑO Y MES
    # ----------------------------------------------------------

    auditorias_errores = auditorias


    if anio_digitador:

        auditorias_errores = (
            auditorias_errores
            .filter(
                fecha_operacion__year=anio_digitador
            )
        )


    if mes_digitador:

        auditorias_errores = (
            auditorias_errores
            .filter(
                fecha_operacion__month=mes_digitador
            )
        )


    # ----------------------------------------------------------
    # SOLO AUDITORÍAS QUE NO CUMPLEN
    # ----------------------------------------------------------

    auditorias_errores = (
        auditorias_errores
        .filter(
            resultado_auditoria="no_cumple"
        )
        .exclude(
            nombre_tecnico__isnull=True
        )
        .exclude(
            nombre_tecnico__exact=""
        )
        .exclude(
            hallazgo__isnull=True
        )
        .exclude(
            hallazgo__exact=""
        )
    )


    # ----------------------------------------------------------
    # OBTENER TÉCNICO + HALLAZGO
    # ----------------------------------------------------------

    datos_errores = list(

        auditorias_errores
        .values(
            "nombre_tecnico",
            "hallazgo"
        )
        .annotate(
            total=Count("id")
        )
    )


    # ----------------------------------------------------------
    # AGRUPAR POR TÉCNICO
    # ----------------------------------------------------------

    errores_por_tecnico = {}


    for dato in datos_errores:

        tecnico = dato["nombre_tecnico"]
        hallazgo = dato["hallazgo"]
        total = dato["total"]

        if tecnico not in errores_por_tecnico:

            errores_por_tecnico[tecnico] = []

        errores_por_tecnico[tecnico].append({

            "hallazgo": hallazgo,

            "total": total

        })


    # ----------------------------------------------------------
    # CALCULAR TOTAL DE CADA TÉCNICO
    # ----------------------------------------------------------

    totales_tecnicos = {}


    for tecnico, hallazgos in errores_por_tecnico.items():

        totales_tecnicos[tecnico] = sum(

            item["total"]

            for item in hallazgos

        )


    # ----------------------------------------------------------
    # ORDENAR TÉCNICOS DE MAYOR A MENOR
    # ----------------------------------------------------------

    tecnicos_ordenados = sorted(

        errores_por_tecnico.keys(),

        key=lambda tecnico:
            totales_tecnicos[tecnico],

        reverse=True

    )


    # ----------------------------------------------------------
    # UBICACIÓN
    # ----------------------------------------------------------

    columna = 10


    # ----------------------------------------------------------
    # TÍTULO
    # ----------------------------------------------------------

    crear_titulo(

        ws,

        1,

        columna,

        columna + 1,

        "ERRORES COMETIDOS POR LOS TÉCNICOS"

    )


    # ----------------------------------------------------------
    # INFORMACIÓN DE FILTROS
    # ----------------------------------------------------------

    # MES

    celda = ws.cell(
        2,
        columna
    )

    celda.value = "MES"

    celda.font = Font(
        bold=True
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=COLOR_HEADER
    )

    celda.border = borde


    celda = ws.cell(
        2,
        columna + 1
    )


    meses = {

        "1": "ENERO",
        "2": "FEBRERO",
        "3": "MARZO",
        "4": "ABRIL",
        "5": "MAYO",
        "6": "JUNIO",
        "7": "JULIO",
        "8": "AGOSTO",
        "9": "SEPTIEMBRE",
        "10": "OCTUBRE",
        "11": "NOVIEMBRE",
        "12": "DICIEMBRE"

    }


    if mes_digitador:

        celda.value = meses.get(
            str(mes_digitador),
            str(mes_digitador)
        )

    else:

        celda.value = "(Todos)"


    celda.border = borde


    # ----------------------------------------------------------
    # FECHA / AÑO
    # ----------------------------------------------------------

    celda = ws.cell(
        3,
        columna
    )

    celda.value = "FECHA"

    celda.font = Font(
        bold=True
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=COLOR_HEADER
    )

    celda.border = borde


    celda = ws.cell(
        3,
        columna + 1
    )


    if anio_digitador:

        celda.value = str(anio_digitador)

    else:

        celda.value = "(Todas)"


    celda.border = borde


    # ----------------------------------------------------------
    # RESULTADO
    # ----------------------------------------------------------

    celda = ws.cell(
        4,
        columna
    )

    celda.value = "Resultado Auditoria"

    celda.font = Font(
        bold=True
    )

    celda.fill = PatternFill(
        "solid",
        fgColor=COLOR_HEADER
    )

    celda.border = borde


    celda = ws.cell(
        4,
        columna + 1
    )

    celda.value = "No cumple"

    celda.border = borde


    # ----------------------------------------------------------
    # ENCABEZADOS
    # ----------------------------------------------------------

    fila_encabezado_errores = 6


    crear_encabezados(

        ws,

        fila_encabezado_errores,

        columna,

        [
            "Etiquetas de fila",
            "Cuenta de Resultado Auditoria"
        ]

    )


    # ----------------------------------------------------------
    # ESCRIBIR INFORMACIÓN
    # ----------------------------------------------------------

    fila_actual = 7


    total_general_errores = 0


    for tecnico in tecnicos_ordenados:

        # ======================================================
        # TOTAL DEL TÉCNICO
        # ======================================================

        total_tecnico = totales_tecnicos[
            tecnico
        ]


        total_general_errores += total_tecnico


        # ------------------------------------------------------
        # FILA DEL TÉCNICO
        # ------------------------------------------------------

        celda_tecnico = ws.cell(

            fila_actual,

            columna

        )

        celda_tecnico.value = tecnico

        celda_tecnico.font = Font(

            bold=True,

            color="000000"

        )

        celda_tecnico.fill = PatternFill(

            "solid",

            fgColor="D9EAF7"

        )

        celda_tecnico.border = borde

        celda_tecnico.alignment = Alignment(

            vertical="center"

        )


        celda_total = ws.cell(

            fila_actual,

            columna + 1

        )

        celda_total.value = total_tecnico

        celda_total.font = Font(

            bold=True,

            color="000000"

        )

        celda_total.fill = PatternFill(

            "solid",

            fgColor="D9EAF7"

        )

        celda_total.border = borde

        celda_total.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )


        fila_actual += 1


        # ======================================================
        # ORDENAR HALLAZGOS DEL TÉCNICO
        # ======================================================

        hallazgos_ordenados = sorted(

            errores_por_tecnico[tecnico],

            key=lambda item:
                item["total"],

            reverse=True

        )


        # ======================================================
        # ESCRIBIR HALLAZGOS
        # ======================================================

        for item in hallazgos_ordenados:

            celda_hallazgo = ws.cell(

                fila_actual,

                columna

            )

            celda_hallazgo.value = (
                item["hallazgo"]
            )

            celda_hallazgo.border = borde

            celda_hallazgo.alignment = Alignment(

                vertical="center",

                indent=1

            )


            celda_cantidad = ws.cell(

                fila_actual,

                columna + 1

            )

            celda_cantidad.value = (
                item["total"]
            )

            celda_cantidad.border = borde

            celda_cantidad.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )


            fila_actual += 1


    # ----------------------------------------------------------
    # TOTAL GENERAL
    # ----------------------------------------------------------

    crear_total(

        ws,

        fila_actual,

        columna,

        [

            "Total general",

            total_general_errores

        ]

    )


    # ----------------------------------------------------------
    # GUARDAR ÚLTIMA FILA
    # ----------------------------------------------------------

    ultima_fila_errores = fila_actual


    # ----------------------------------------------------------
    # IMPORTANTE:
    # NO CREAR Table() AQUÍ
    # ----------------------------------------------------------
    #
    # Esta sección NO utiliza:
    #
    # crear_tabla(...)
    #
    # porque Excel estaba reparando/eliminando las tablas.
    #
    # La información queda como celdas normales con formato.
    # ----------------------------------------------------------


    # ----------------------------------------------------------
    # ANCHO DE COLUMNAS
    # ----------------------------------------------------------

    ajustar_ancho(

        columna,

        columna + 1,

        50

    )


    # ==========================================================
    # 5. CANTIDAD DE HALLAZGOS ENCONTRADOS
    # ==========================================================

    datos = (

        auditorias_hallazgos

        .filter(

            resultado_auditoria="no_cumple"

        )

        .exclude(

            hallazgo__isnull=True

        )

        .exclude(

            hallazgo__exact=""

        )

        .values(

            "hallazgo"

        )

        .annotate(

            total=Count("id")

        )

        .order_by(

            "-total"

        )

    )


    datos = list(datos)


    columna = 14


    crear_titulo(

        ws,

        1,

        columna,

        columna + 1,

        "CANTIDAD DE HALLAZGOS ENCONTRADOS"

    )


    crear_encabezados(

        ws,

        2,

        columna,

        [

            "Hallazgo",

            "Total"

        ]

    )


    filas = []

    total_hallazgos = 0


    for dato in datos:

        filas.append([

            dato["hallazgo"],

            dato["total"]

        ])


        total_hallazgos += dato["total"]


    ultima_fila_hallazgos = escribir_datos(

        ws,

        3,

        columna,

        filas

    )


    # ----------------------------------------------------------
    # TOTAL FUERA DE LA TABLA
    # ----------------------------------------------------------

    crear_total(

        ws,

        ultima_fila_hallazgos + 1,

        columna,

        [

            "Total general",

            total_hallazgos

        ]

    )


    # ----------------------------------------------------------
    # TABLA SOLO CON ENCABEZADO + DATOS
    # ----------------------------------------------------------

    crear_tabla(

        ws,

        "TablaCantidadHallazgos",

        2,

        columna,

        ultima_fila_hallazgos,

        columna + 1

    )


    ajustar_ancho(

        columna,

        columna + 1,

        45

    )


    # ==========================================================
    # 6. AUDITORÍAS POR DIGITADOR
    # ==========================================================

    datos = list(

        auditorias_digitador

        .exclude(

            nombre_auditor__isnull=True

        )

        .exclude(

            nombre_auditor__exact=""

        )

        .values(

            "fecha_operacion",

            "nombre_auditor"

        )

        .annotate(

            total=Count("id")

        )

        .order_by(

            "fecha_operacion",

            "nombre_auditor"

        )

    )


    digitadores = sorted(

        set(

            dato["nombre_auditor"]

            for dato in datos

        )

    )


    columna = 17


    ultima_columna = (

        columna +

        len(digitadores) +

        1

    )


    crear_titulo(

        ws,

        1,

        columna,

        ultima_columna,

        "AUDITORÍAS POR DIGITADOR"

    )


    encabezado = [

        "Fecha"

    ]


    encabezado.extend(

        digitadores

    )


    encabezado.append(

        "Total general"

    )


    crear_encabezados(

        ws,

        2,

        columna,

        encabezado

    )


    mapa = {}


    for dato in datos:

        fecha = dato[
            "fecha_operacion"
        ]

        digitador = dato[
            "nombre_auditor"
        ]

        cantidad = dato[
            "total"
        ]


        if fecha not in mapa:

            mapa[fecha] = {}


        mapa[fecha][
            digitador
        ] = cantidad


    totales = {

        digitador: 0

        for digitador in digitadores

    }


    total_general_digitadores = 0


    fila_actual = 3


    for fecha in sorted(
        mapa.keys()
    ):

        fila = [

            fecha

        ]


        total_dia = 0


        for digitador in digitadores:

            cantidad = (

                mapa[fecha]

                .get(

                    digitador,

                    0

                )

            )


            fila.append(

                cantidad

            )


            totales[
                digitador
            ] += cantidad


            total_dia += cantidad


        fila.append(

            total_dia

        )


        total_general_digitadores += (

            total_dia

        )


        for i, valor in enumerate(
            fila
        ):

            celda = ws.cell(

                fila_actual,

                columna + i

            )


            celda.value = valor


            celda.border = borde


            celda.alignment = Alignment(

                vertical="center"

            )


            if hasattr(

                valor,

                "strftime"

            ):

                celda.number_format = (
                    "dd/mm/yyyy"
                )


        fila_actual += 1


    # ----------------------------------------------------------
    # TOTAL DIGITADORES
    # ----------------------------------------------------------

    fila_total = [

        "Total general"

    ]


    for digitador in digitadores:

        fila_total.append(

            totales[digitador]

        )


    fila_total.append(

        total_general_digitadores

    )


    crear_total(

        ws,

        fila_actual,

        columna,

        fila_total

    )


    # ----------------------------------------------------------
    # TABLA DIGITADORES
    # ----------------------------------------------------------

    crear_tabla(

        ws,

        "TablaAuditoriasDigitador",

        2,

        columna,

        fila_actual - 1,

        ultima_columna

    )


    ajustar_ancho(

        columna,

        ultima_columna,

        30

    )


    # ==========================================================
    # 7. OPERACIONES
    # ==========================================================

    fila_operaciones = max(

        3,

        max(

            ultima_fila_totales,

            ultima_fila_no_cumplen,

            ultima_fila_tecnicos,

            ultima_fila_errores,

            ultima_fila_hallazgos

        ) + 4

    )


    columna = 1


    crear_titulo(

        ws,

        fila_operaciones,

        columna,

        columna + 1,

        "OPERACIONES"

    )


    crear_encabezados(

        ws,

        fila_operaciones + 1,

        columna,

        [

            "Tipo operación",

            "Cantidad"

        ]

    )


    operaciones = [

        (
            "Suspensiones",
            "DC00"
        ),

        (
            "Reconexiones",
            "RC00"
        ),

        (
            "ZVCL",
            "ZVCL"
        )

    ]


    filas = []

    total_operaciones = 0


    for nombre, codigo in operaciones:

        cantidad = auditorias.filter(

            tipo_operacion=codigo

        ).count()


        filas.append([

            nombre,

            cantidad

        ])


        total_operaciones += cantidad


    ultima_fila_operaciones = (

        escribir_datos(

            ws,

            fila_operaciones + 2,

            columna,

            filas

        )

    )


    crear_total(

        ws,

        ultima_fila_operaciones + 1,

        columna,

        [

            "Total general",

            total_operaciones

        ]

    )


    crear_tabla(

        ws,

        "TablaOperaciones",

        fila_operaciones + 1,

        columna,

        ultima_fila_operaciones,

        columna + 1

    )


    ajustar_ancho(

        columna,

        columna + 1

    )


    # ==========================================================
    # 8. RESUMEN GENERAL
    # ==========================================================

    columna = 4


    crear_titulo(

        ws,

        fila_operaciones,

        columna,

        columna + 1,

        "RESUMEN GENERAL"

    )


    crear_encabezados(

        ws,

        fila_operaciones + 1,

        columna,

        [

            "Estadística",

            "Cantidad"

        ]

    )


    resumen = [

        (

            "Total auditorías",

            auditorias.count()

        ),

        (

            "No cumplen",

            total_no_cumplen

        ),

        (

            "Suspensiones",

            auditorias.filter(

                tipo_operacion="DC00"

            ).count()

        ),

        (

            "Reconexiones",

            auditorias.filter(

                tipo_operacion="RC00"

            ).count()

        ),

        (

            "ZVCL",

            auditorias.filter(

                tipo_operacion="ZVCL"

            ).count()

        ),

        (

            "Técnicos",

            auditorias

            .values(

                "nombre_tecnico"

            )

            .distinct()

            .count()

        ),

        (

            "Digitadores",

            auditorias

            .exclude(

                nombre_auditor__isnull=True

            )

            .exclude(

                nombre_auditor__exact=""

            )

            .values(

                "nombre_auditor"

            )

            .distinct()

            .count()

        )

    ]


    filas = []


    for nombre, cantidad in resumen:

        filas.append([

            nombre,

            cantidad

        ])


    ultima_fila_resumen = escribir_datos(

        ws,

        fila_operaciones + 2,

        columna,

        filas

    )


    crear_total(

        ws,

        ultima_fila_resumen + 1,

        columna,

        [

            "Total",

            auditorias.count()

        ]

    )


    crear_tabla(

        ws,

        "TablaResumen",

        fila_operaciones + 1,

        columna,

        ultima_fila_resumen,

        columna + 1

    )


    ajustar_ancho(

        columna,

        columna + 1

    )


    # ==========================================================
    # FORMATO FINAL
    # ==========================================================

    for fila in range(

        1,

        ws.max_row + 1

    ):

        if ws.row_dimensions[
            fila
        ].height is None:

            ws.row_dimensions[
                fila
            ].height = 18


    # ==========================================================
    # ALINEACIÓN DE NÚMEROS
    # ==========================================================

    for fila in ws.iter_rows():

        for celda in fila:

            if isinstance(

                celda.value,

                (int, float)

            ):

                celda.alignment = Alignment(

                    horizontal="center",

                    vertical="center"

                )


    # ==========================================================
    # CONFIGURACIÓN DE IMPRESIÓN
    # ==========================================================

    ws.page_setup.orientation = (
        "landscape"
    )


    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 0


    ws.sheet_properties.pageSetUpPr.fitToPage = True


    ws.print_title_rows = "1:2"


    # ==========================================================
    # DESCARGAR EXCEL
    # ==========================================================

    response = HttpResponse(

        content_type=(

            "application/vnd.openxmlformats-officedocument."

            "spreadsheetml.sheet"

        )

    )


    response["Content-Disposition"] = (

        'attachment; '

        'filename="Estadisticas_Auditorias.xlsx"'

    )


    wb.save(
        response
    )


    return response