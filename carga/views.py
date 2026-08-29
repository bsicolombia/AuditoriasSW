from datetime import datetime, date
import json
import re

from django.apps import apps
from django.contrib import messages
from django.db import transaction
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.utils import timezone

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from xhtml2pdf import pisa

from .forms import (
    CsvUploadForm,
    AuditoriaForm,
    TecnicoForm,
    TecnicoCargaForm,
)

from auditorias.models import Auditoria
from .models import Tecnicos

# ==========================================================
# CARGA
# ==========================================================

def carga(request):

    context = {

        # Formulario auditorías
        "form": CsvUploadForm(),

        # Formulario técnicos
        "tecnico_form": TecnicoCargaForm(),

        # ==================================================
        # ÚLTIMA CARGA DE AUDITORÍAS
        # ==================================================

        "ultima_carga": request.session.get(
            "ultima_carga_auditorias"
        ),

        # ==================================================
        # ÚLTIMA CARGA DE TÉCNICOS
        # ==================================================

        "ultima_carga_tecnicos": request.session.get(
            "ultima_carga_tecnicos"
        ),

        # ==================================================
        # REPORTE DE TÉCNICOS
        # ==================================================

        "tecnico_report_generated": request.session.get(
            "tecnico_report_generated",
            False,
        ),

        "tecnico_total_rows": request.session.get(
            "tecnico_total_rows",
            0,
        ),

        "tecnico_created_count": request.session.get(
            "tecnico_created_count",
            0,
        ),

        "tecnico_updated_count": request.session.get(
            "tecnico_updated_count",
            0,
        ),

        "tecnico_existing_count": request.session.get(
            "tecnico_existing_count",
            0,
        ),

        "tecnico_error_count": request.session.get(
            "tecnico_error_count",
            0,
        ),

        "tecnico_errors": request.session.get(
            "tecnico_errors",
            [],
        ),
    }

    return render(
        request,
        "carga/carga.html",
        context,
    )


# ==========================================================
# CONVERTIR TEXTO
# ==========================================================

def convertir_texto(value):

    if value is None:
        return ""

    if isinstance(value, float):

        if value.is_integer():
            return str(int(value))

        return str(value)

    return str(value).strip()


# ==========================================================
# CONVERTIR FECHA
# ==========================================================

def convertir_fecha(value):

    if value is None:
        return None

    # ------------------------------------------------------
    # DATETIME
    # ------------------------------------------------------

    if isinstance(value, datetime):
        return value.date()

    # ------------------------------------------------------
    # DATE
    # ------------------------------------------------------

    if isinstance(value, date):
        return value

    # ------------------------------------------------------
    # TEXTO
    # ------------------------------------------------------

    if isinstance(value, str):

        value = value.strip()

        if not value:
            return None

        formatos = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%B %d, %Y",
        ]

        for formato in formatos:

            try:

                return datetime.strptime(
                    value,
                    formato
                ).date()

            except ValueError:
                continue

        return None

    return None


# ==========================================================
# VALIDAR FECHA DE AUDITORÍA
# ==========================================================

def validar_fecha_auditoria(value):

    """
    Valida que la fecha:

    - Sea una fecha válida.
    - Tenga año de 4 dígitos.
    - No sea anterior a 2025.
    - No sea superior a la fecha actual.
    - No tenga años como 0026.
    - No tenga años futuros.
    """

    if value is None:

        return (
            False,
            "La fecha es obligatoria."
        )

    # ------------------------------------------------------
    # CONVERTIR A DATE
    # ------------------------------------------------------

    if isinstance(value, datetime):

        fecha = value.date()

    elif isinstance(value, date):

        fecha = value

    else:

        return (
            False,
            "La fecha no tiene un formato válido."
        )

    # ------------------------------------------------------
    # FECHA ACTUAL
    # ------------------------------------------------------

    hoy = date.today()

    # ------------------------------------------------------
    # AÑO MENOR A 1000
    # ------------------------------------------------------

    if fecha.year < 1000:

        return (
            False,
            (
                f"El año {fecha.year:04d} no es válido. "
                "El año debe tener 4 dígitos."
            )
        )

    # ------------------------------------------------------
    # AÑO MÍNIMO
    # ------------------------------------------------------

    año_minimo = 2025

    if fecha.year < año_minimo:

        return (
            False,
            (
                f"El año {fecha.year} no es válido. "
                f"La fecha debe corresponder al año "
                f"{año_minimo} o posterior."
            )
        )

    # ------------------------------------------------------
    # FECHA FUTURA
    # ------------------------------------------------------

    if fecha > hoy:

        return (
            False,
            (
                f"La fecha {fecha.strftime('%d/%m/%Y')} "
                f"es superior a la fecha actual "
                f"({hoy.strftime('%d/%m/%Y')})."
            )
        )

    # ------------------------------------------------------
    # AÑO FUTURO
    # ------------------------------------------------------

    if fecha.year > hoy.year:

        return (
            False,
            (
                f"El año {fecha.year} no es válido. "
                f"No se permiten años posteriores "
                f"a {hoy.year}."
            )
        )

    return (
        True,
        ""
    )


# ==========================================================
# VALIDAR TEXTO DE FECHA
# ==========================================================

def validar_texto_fecha(value):

    """
    Valida específicamente que un texto tenga año de 4 dígitos.

    Ejemplos inválidos:

        25/08/26
        25/08/0026
        2026/08/25?  -> se valida según formatos permitidos

    Ejemplos válidos:

        25/08/2026
        2026-08-25
        08/25/2026
    """

    if not isinstance(value, str):

        return True

    value = value.strip()

    if not value:
        return True

    numeros = re.findall(
        r"\d+",
        value
    )

    # ------------------------------------------------------
    # Detectar años escritos con cantidad incorrecta
    # ------------------------------------------------------

    for numero in numeros:

        numero_int = int(numero)

        # Si tiene 2 dígitos y parece año
        if len(numero) == 2:

            # Ej: 25/08/26
            if numero_int <= 99:

                return False

        # Año con menos de 4 dígitos
        elif len(numero) == 3:

            return False

        # Año escrito como 0026
        elif len(numero) == 4:

            # Si empieza por 0, no es un año válido
            if numero.startswith("0"):

                return False

    return True


# ==========================================================
# NORMALIZAR RESULTADO
# ==========================================================

def normalizar_resultado(value):

    if value is None:
        return ""

    valor = str(value).strip().lower()

    equivalencias = {

        "cumple":
            "cumple",

        "no cumple":
            "no_cumple",

        "no_cumple":
            "no_cumple",

        "nocumple":
            "no_cumple",
    }

    return equivalencias.get(
        valor,
        valor
    )


# ==========================================================
# NORMALIZAR TIPO DE HALLAZGO
# ==========================================================

def normalizar_tipo_hallazgo(value):

    if value is None:
        return ""

    valor = str(value).strip().lower()

    equivalencias = {

        "alto":
            "alto",

        "medio":
            "medio",

        "bajo":
            "bajo",
    }

    return equivalencias.get(
        valor,
        valor
    )


# ==========================================================
# GENERAR FIRMA DEL REGISTRO
# ==========================================================

def generar_firma_registro(data):

    campos = [

        "fecha",
        "nombre_auditor",
        "numero_cedula",
        "aplicativo",
        "fecha_operacion",
        "nombre_tecnico",
        "numero_cuenta_contrato",
        "numero_orden",
        "tipo_operacion",
        "resultado_auditoria",
        "observacion",
        "tipo_hallazgo",
        "hallazgo",
    ]

    valores = []

    for campo in campos:

        valor = data.get(
            campo,
            ""
        )

        if isinstance(valor, date):

            valor = valor.isoformat()

        elif valor is None:

            valor = ""

        else:

            valor = str(
                valor
            ).strip().lower()

        valores.append(
            valor
        )

    return tuple(
        valores
    )


# ==========================================================
# OBTENER HALLAZGO
# ==========================================================

def obtener_hallazgo(
    tipo_hallazgo,
    hallazgo_alto,
    hallazgo_medio,
    hallazgo_bajo
):

    tipo_hallazgo = (
        tipo_hallazgo or ""
    ).strip().lower()

    if tipo_hallazgo == "alto":

        return convertir_texto(
            hallazgo_alto
        )

    if tipo_hallazgo == "medio":

        return convertir_texto(
            hallazgo_medio
        )

    if tipo_hallazgo == "bajo":

        return convertir_texto(
            hallazgo_bajo
        )

    return ""


# ==========================================================
# PREPARAR DATOS PARA JSON
# ==========================================================

def preparar_para_json(obj):

    if isinstance(obj, dict):

        return {
            str(key): preparar_para_json(value)
            for key, value in obj.items()
        }

    if isinstance(obj, (list, tuple)):

        return [
            preparar_para_json(value)
            for value in obj
        ]

    if isinstance(obj, (datetime, date)):

        return obj.isoformat()

    if obj is None:

        return ""

    return obj


# ==========================================================
# CREAR REGISTRO DE ERROR
# ==========================================================

def crear_error(
    row_number,
    form_data,
    campo,
    mensaje
):

    return {
        "row": row_number,

        # IMPORTANTE:
        # aquí se conserva TODA la fila
        "data": form_data.copy(),

        "errors": {
            campo: mensaje
        },
    }


# ==========================================================
# PROCESAR EXCEL
# ==========================================================

def auditoria_crear(request):

    # ------------------------------------------------------
    # SOLO POST
    # ------------------------------------------------------

    if request.method != "POST":

        form = CsvUploadForm()

        return render(
            request,
            "carga/carga.html",
            {
                "form": form,
            }
        )

    # ------------------------------------------------------
    # FORMULARIO
    # ------------------------------------------------------

    upload_form = CsvUploadForm(
        request.POST,
        request.FILES
    )

    if not upload_form.is_valid():

        return render(
            request,
            "carga/carga.html",
            {
                "form": upload_form,
                "report_generated": False,
            }
        )

    # ------------------------------------------------------
    # ARCHIVO
    # ------------------------------------------------------

    excel_file = request.FILES.get(
        "csv_file"
    )

    if not excel_file:

        return render(
            request,
            "carga/carga.html",
            {
                "form": upload_form,
                "error": (
                    "No se recibió ningún archivo."
                ),
            }
        )

    # ------------------------------------------------------
    # EXTENSIÓN
    # ------------------------------------------------------

    if not excel_file.name.lower().endswith(".xlsx"):

        return render(
            request,
            "carga/carga.html",
            {
                "form": upload_form,
                "error": (
                    "El archivo debe ser un Excel "
                    "con extensión .xlsx."
                ),
            }
        )

    # ------------------------------------------------------
    # ABRIR EXCEL
    # ------------------------------------------------------

    try:

        workbook = load_workbook(
            excel_file,
            data_only=True
        )

        worksheet = workbook.active

    except Exception as e:

        return render(
            request,
            "carga/carga.html",
            {
                "form": upload_form,
                "error": (
                    "No se pudo abrir el archivo Excel: "
                    f"{e}"
                ),
            }
        )

    # ------------------------------------------------------
    # FILAS
    # ------------------------------------------------------

    rows = worksheet.iter_rows(
        values_only=True
    )

    # ------------------------------------------------------
    # ENCABEZADOS
    # ------------------------------------------------------

    try:

        headers = next(rows)

    except StopIteration:

        return render(
            request,
            "carga/carga.html",
            {
                "form": upload_form,
                "error": (
                    "El archivo Excel está vacío."
                ),
            }
        )

    # ------------------------------------------------------
    # NORMALIZAR ENCABEZADOS
    # ------------------------------------------------------

    headers = [

        str(header).strip().lower()
        if header is not None
        else ""

        for header in headers
    ]

    # ======================================================
    # MAPEO DE COLUMNAS
    # ======================================================

    column_mapping = {

        "marca temporal":
            "fecha",

        "fecha":
            "fecha",

        "nombre auditor":
            "nombre_auditor",

        "numero de cedula":
            "numero_cedula",

        "número de cédula":
            "numero_cedula",

        "numero cedula":
            "numero_cedula",

        "aplicativo":
            "aplicativo",

        "fecha operación":
            "fecha_operacion",

        "fecha operacion":
            "fecha_operacion",

        "nombres y apellidos técnico":
            "nombre_tecnico",

        "nombres y apellidos tecnico":
            "nombre_tecnico",

        "nombre técnico":
            "nombre_tecnico",

        "nombre tecnico":
            "nombre_tecnico",

        "número de cuenta contrato":
            "numero_cuenta_contrato",

        "numero de cuenta contrato":
            "numero_cuenta_contrato",

        "número de orden":
            "numero_orden",

        "numero de orden":
            "numero_orden",

        "tipo de operación":
            "tipo_operacion",

        "tipo de operacion":
            "tipo_operacion",

        "resultado auditoria":
            "resultado_auditoria",

        "resultado auditoría":
            "resultado_auditoria",

        "observacion":
            "observacion",

        "observación":
            "observacion",

        "tipo de hallazgo":
            "tipo_hallazgo",

        "hallazgo":
            "hallazgo",

        "columna 12":
            "hallazgo_alto",

        "columna 13":
            "hallazgo_medio",

        "columna 14":
            "hallazgo_bajo",
    }

    headers = [

        column_mapping.get(
            header,
            None
        )

        for header in headers
    ]

    # ======================================================
    # CAMPOS DEL MODELO
    # ======================================================

    campos_modelo = {

        "fecha",

        "nombre_auditor",

        "numero_cedula",

        "aplicativo",

        "fecha_operacion",

        "nombre_tecnico",

        "numero_cuenta_contrato",

        "numero_orden",

        "tipo_operacion",

        "resultado_auditoria",

        "observacion",

        "tipo_hallazgo",

        "hallazgo",
    }

    # ======================================================
    # VARIABLES
    # ======================================================

    successful_records = []

    error_records = []

    auditorias_creadas = []

    # Registros que ya estaban correctamente guardados
    registros_existentes = []

    # Firmas de registros NUEVOS del archivo
    firmas_excel = set()

    # Órdenes NUEVAS del archivo
    ordenes_excel = set()

    total_rows = 0

    # ======================================================
    # PROCESAR FILAS
    # ======================================================

    for row_number, row in enumerate(
        rows,
        start=2
    ):

        # --------------------------------------------------
        # IGNORAR FILA VACÍA
        # --------------------------------------------------

        if not any(
            value is not None
            and str(value).strip() != ""
            for value in row
        ):

            continue

        total_rows += 1

        # --------------------------------------------------
        # HALLAZGOS
        # --------------------------------------------------

        hallazgo_alto = ""

        hallazgo_medio = ""

        hallazgo_bajo = ""

        # --------------------------------------------------
        # DATOS
        # --------------------------------------------------

        form_data = {}

        # --------------------------------------------------
        # ERRORES DE FECHA
        # --------------------------------------------------

        errores_fecha = []

        # ==================================================
        # LEER TODA LA FILA
        # ==================================================

        for header, value in zip(
            headers,
            row
        ):

            if not header:
                continue

            # ----------------------------------------------
            # HALLAZGO ALTO
            # ----------------------------------------------

            if header == "hallazgo_alto":

                hallazgo_alto = convertir_texto(
                    value
                )

                continue

            # ----------------------------------------------
            # HALLAZGO MEDIO
            # ----------------------------------------------

            if header == "hallazgo_medio":

                hallazgo_medio = convertir_texto(
                    value
                )

                continue

            # ----------------------------------------------
            # HALLAZGO BAJO
            # ----------------------------------------------

            if header == "hallazgo_bajo":

                hallazgo_bajo = convertir_texto(
                    value
                )

                continue

            # ----------------------------------------------
            # COLUMNA DESCONOCIDA
            # ----------------------------------------------

            if header not in campos_modelo:

                continue

            # ==================================================
            # FECHAS
            # ==================================================

            if header in [
                "fecha",
                "fecha_operacion",
            ]:

                valor_original = value

                # ------------------------------------------
                # VALIDAR TEXTO
                # ------------------------------------------

                if isinstance(
                    valor_original,
                    str
                ):

                    if not validar_texto_fecha(
                        valor_original
                    ):

                        errores_fecha.append(
                            (
                                header,
                                (
                                    "La fecha no es válida. "
                                    "El año debe tener "
                                    "exactamente 4 dígitos. "
                                    "Ejemplo: 25/08/2026."
                                )
                            )
                        )

                        form_data[header] = (
                            convertir_texto(
                                valor_original
                            )
                        )

                        continue

                # ------------------------------------------
                # CONVERTIR
                # ------------------------------------------

                fecha_convertida = convertir_fecha(
                    valor_original
                )

                if fecha_convertida is None:

                    errores_fecha.append(
                        (
                            header,
                            (
                                "La fecha no es válida. "
                                "Debe utilizar un formato "
                                "válido y un año de 4 dígitos."
                            )
                        )
                    )

                    form_data[header] = (
                        convertir_texto(
                            valor_original
                        )
                    )

                    continue

                # ------------------------------------------
                # VALIDAR FECHA
                # ------------------------------------------

                fecha_valida, mensaje_fecha = (
                    validar_fecha_auditoria(
                        fecha_convertida
                    )
                )

                if not fecha_valida:

                    errores_fecha.append(
                        (
                            header,
                            mensaje_fecha
                        )
                    )

                form_data[header] = (
                    fecha_convertida
                )

            # ==================================================
            # NÚMEROS COMO TEXTO
            # ==================================================

            elif header in [

                "numero_cedula",

                "numero_cuenta_contrato",

                "numero_orden",
            ]:

                form_data[header] = (
                    convertir_texto(
                        value
                    )
                )

            # ==================================================
            # RESULTADO
            # ==================================================

            elif header == "resultado_auditoria":

                form_data[header] = (
                    normalizar_resultado(
                        value
                    )
                )

            # ==================================================
            # TIPO DE HALLAZGO
            # ==================================================

            elif header == "tipo_hallazgo":

                form_data[header] = (
                    normalizar_tipo_hallazgo(
                        value
                    )
                )

            # ==================================================
            # TEXTOS
            # ==================================================

            elif header in [

                "nombre_auditor",

                "aplicativo",

                "nombre_tecnico",

                "tipo_operacion",

                "observacion",

                "hallazgo",
            ]:

                form_data[header] = (
                    convertir_texto(
                        value
                    )
                )

        # ==================================================
        # VALORES POR DEFECTO
        # ==================================================

        form_data.setdefault(
            "observacion",
            ""
        )

        form_data.setdefault(
            "tipo_hallazgo",
            ""
        )

        # ==================================================
        # OBTENER HALLAZGO
        # ==================================================

        form_data["hallazgo"] = obtener_hallazgo(
            form_data.get(
                "tipo_hallazgo",
                ""
            ),
            hallazgo_alto,
            hallazgo_medio,
            hallazgo_bajo
        )

        # ==================================================
        # VALIDAR FECHAS
        # ==================================================

        if errores_fecha:

            for campo, mensaje in errores_fecha:

                error_records.append(
                    crear_error(
                        row_number,
                        form_data,
                        campo,
                        mensaje
                    )
                )

            # IMPORTANTE:
            # NO se registra como procesado.
            # En la siguiente carga volverá a intentarse.
            continue

        # ==================================================
        # VALIDAR FORMULARIO
        # ==================================================

        row_form = AuditoriaForm(
            data=form_data
        )

        if not row_form.is_valid():

            errors = {}

            for field, error_list in row_form.errors.items():

                errors[field] = ", ".join(
                    str(error)
                    for error in error_list
                )

            error_records.append(
                {
                    "row": row_number,

                    "data": form_data.copy(),

                    "errors": errors,
                }
            )

            # IMPORTANTE:
            # No se guarda.
            # Por eso, en la próxima carga
            # volverá a aparecer el error.
            continue

        # ==================================================
        # NÚMERO DE ORDEN
        # ==================================================

        numero_orden = str(
            form_data.get(
                "numero_orden",
                ""
            )
        ).strip()

        # ==================================================
        # COMPROBAR SI YA EXISTE EN BD
        # ==================================================

        try:

            registro_existente = None

            if numero_orden:

                registro_existente = (
                    Auditoria.objects
                    .filter(
                        numero_orden=numero_orden
                    )
                    .first()
                )

        except Exception as e:

            error_records.append(
                {
                    "row": row_number,

                    "data": form_data.copy(),

                    "errors": {
                        "Base de datos": (
                            "No fue posible comprobar "
                            "si la auditoría ya existe: "
                            f"{e}"
                        )
                    },
                }
            )

            continue

        # ==================================================
        # YA EXISTE EN BD
        # ==================================================

        if registro_existente:

            # NO es error.
            #
            # Significa que esta auditoría ya fue
            # procesada correctamente en una carga
            # anterior.
            registros_existentes.append(
                {
                    "row": row_number,

                    "data": form_data.copy(),
                }
            )

            continue

        # ==================================================
        # DESDE AQUÍ EL REGISTRO ES NUEVO
        # ==================================================

        # ==================================================
        # DUPLICADO EXACTO ENTRE REGISTROS NUEVOS
        # ==================================================

        firma = generar_firma_registro(
            form_data
        )

        if firma in firmas_excel:

            error_records.append(
                {
                    "row": row_number,

                    "data": form_data.copy(),

                    "errors": {
                        "Registro duplicado": (
                            "Esta auditoría aparece "
                            "más de una vez dentro "
                            "de las nuevas auditorías "
                            "del archivo."
                        )
                    },
                }
            )

            continue

        firmas_excel.add(
            firma
        )

        # ==================================================
        # DUPLICADO DE ORDEN ENTRE REGISTROS NUEVOS
        # ==================================================

        if numero_orden:

            if numero_orden in ordenes_excel:

                error_records.append(
                    {
                        "row": row_number,

                        "data": form_data.copy(),

                        "errors": {
                            "Número de orden": (
                                f"La orden "
                                f"{numero_orden} "
                                "aparece más de una vez "
                                "entre las nuevas "
                                "auditorías del archivo."
                            )
                        },
                    }
                )

                continue

            ordenes_excel.add(
                numero_orden
            )

        # ==================================================
        # PREPARAR AUDITORÍA PARA INSERTAR
        # ==================================================

        try:

            auditoria = row_form.save(
                commit=False
            )

            auditorias_creadas.append(
                auditoria
            )

            successful_records.append(
                {
                    "row": row_number,

                    "data": form_data.copy(),
                }
            )

        except Exception as e:

            error_records.append(
                {
                    "row": row_number,

                    "data": form_data.copy(),

                    "errors": {
                        "Registro": (
                            "No fue posible preparar "
                            "el registro para guardar: "
                            f"{e}"
                        )
                    },
                }
            )

    # ======================================================
    # GUARDAR AUDITORÍAS
    # ======================================================

    if auditorias_creadas:

        try:

            with transaction.atomic():

                Auditoria.objects.bulk_create(
                    auditorias_creadas,
                    batch_size=1000
                )

            # --------------------------------------------------
            # ÚLTIMA CARGA
            # --------------------------------------------------

            request.session[
                "ultima_carga_auditorias"
            ] = (
                timezone.localtime().strftime(
                    "%d/%m/%Y %H:%M:%S"
                )
            )

            request.session.modified = True

        except Exception as e:

            error_records.append(
                {
                    "row": "-",

                    "data": {},

                    "errors": {
                        "Base de datos": (
                            "No fue posible guardar "
                            "las auditorías. "
                            f"Error: {e}"
                        )
                    },
                }
            )

            successful_records = []

            auditorias_creadas = []

    # ======================================================
    # CONTADORES
    # ======================================================

    successful_count = len(successful_records)

    # ======================================================
    # TIPOS DE ERROR ÚNICOS
    # ======================================================

    tipos_error_unicos = set()

    for record in error_records:

        for campo in record.get("errors", {}).keys():

            tipos_error_unicos.add(
                str(campo).strip()
            )

    error_count = len(tipos_error_unicos)

    existing_count = len(
        registros_existentes
    )

    # ======================================================
    # MENSAJES
    # ======================================================

    if successful_count:

        messages.success(
            request,
            (
                "Proceso finalizado correctamente. "
                f"{successful_count} auditorías "
                "_ok."
            )
        )

    if existing_count:

        messages.info(
            request,
            (
                f"{existing_count} auditorías "
                "ya habían sido cargadas anteriormente "
                "y fueron ignoradas."
            )
        )

    if error_count:

        messages.warning(
            request,
            (
                f"{error_count} Auditorias "
                "Presentan Errores por diligenciamiento del Auditor."
            )
        )

    # ======================================================
    # JSON PARA PDF
    # ======================================================

    errores_json_data = preparar_para_json(
        error_records
    )

    errores_json = json.dumps(
        errores_json_data,
        ensure_ascii=False
    )

    # ======================================================
    # ÚLTIMA CARGA
    # ======================================================

    ultima_carga = request.session.get(
        "ultima_carga_auditorias"
    )

    # ======================================================
    # CONTEXTO
    # ======================================================

    context = {

        "form":
            upload_form,

        "successful_count":
            successful_count,

        "error_count":
            error_count,

        "existing_count":
            existing_count,

        "total_rows":
            total_rows,

        "errors_records":
            error_records,

        "successful_records":
            successful_records,

        "existing_records":
            registros_existentes,

        "report_generated":
            True,

        "errores_json":
            errores_json,

        "ultima_carga":
            ultima_carga,
    }

    return render(
        request,
        "carga/carga.html",
        context
    )


def tecnicos_crear(request):

    if request.method != "POST":

        return redirect("carga")


    form = TecnicoCargaForm(request.POST, request.FILES)


    if not form.is_valid():

        context = {

            "form": CsvUploadForm(),

            "tecnico_form": form,

            "ultima_carga": request.session.get(
                "ultima_carga_auditorias"
            ),

            "ultima_carga_tecnicos": request.session.get(
                "ultima_carga_tecnicos"
            ),

            "tecnico_report_generated": False,

        }

        return render(
            request,
            "carga/carga.html",
            context,
        )


    archivo = form.cleaned_data["xlsx_file"]


    # ======================================================
    # VARIABLES DEL REPORTE
    # ======================================================

    total_rows = 0
    created_count = 0
    updated_count = 0
    existing_count = 0
    error_count = 0

    errors = []


    # ======================================================
    # ABRIR EXCEL
    # ======================================================

    try:

        workbook = load_workbook(
            filename=archivo,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

    except Exception as e:

        messages.error(
            request,
            f"No fue posible abrir el archivo Excel: {e}",
        )

        return redirect("carga")


    # ======================================================
    # LEER ENCABEZADOS
    # ======================================================

    try:

        headers = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
        )

    except StopIteration:

        messages.error(
            request,
            "El archivo Excel está vacío.",
        )

        return redirect("carga")


    # ======================================================
    # NORMALIZAR ENCABEZADOS
    # ======================================================

    def normalizar_header(valor):

        if valor is None:
            return ""

        texto = str(valor).strip().lower()

        texto = (
            texto
            .replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
            .replace("ñ", "n")
        )

        texto = re.sub(
            r"[^a-z0-9]+",
            "_",
            texto,
        )

        return texto.strip("_")


    header_map = {}

    for indice, header in enumerate(headers):

        nombre = normalizar_header(header)

        if nombre:

            header_map[nombre] = indice


    # ======================================================
    # COLUMNAS ESPERADAS
    # ======================================================

    columnas_supervisor = [
        "supervisor",
        "nombre_supervisor",
    ]

    columnas_cedula = [
        "tecnico_cedula",
        "cedula",
        "cedula_tecnico",
        "documento",
        "documento_tecnico",
        "numero_cedula",
        "numero_de_cedula",
    ]

    columnas_nombre = [
        "tecnico_apellido_nombres",
        "apellido_nombres",
        "nombre_tecnico",
        "tecnico",
        "nombres_apellidos",
        "apellidos_y_nombres",
        "apellidos_nombres",
    ]

    def buscar_columna(opciones):

        for opcion in opciones:

            if opcion in header_map:
                return header_map[opcion]

        return None


    col_supervisor = buscar_columna(
        columnas_supervisor
    )

    col_cedula = buscar_columna(
        columnas_cedula
    )

    col_nombre = buscar_columna(
        columnas_nombre
    )


    # ======================================================
    # VALIDAR COLUMNAS
    # ======================================================

    columnas_faltantes = []


    if col_supervisor is None:
        columnas_faltantes.append("supervisor")


    if col_cedula is None:
        columnas_faltantes.append(
            "tecnico_cedula / cédula"
        )


    if col_nombre is None:
        columnas_faltantes.append(
            "tecnico_apellido_nombres / nombre técnico"
        )


    if columnas_faltantes:

        workbook.close()

        messages.error(
            request,
            "El Excel no contiene las columnas requeridas: "
            + ", ".join(columnas_faltantes),
        )

        return redirect("carga")


    # ======================================================
    # PROCESAR FILAS
    # ======================================================

    for numero_fila, fila in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        # --------------------------------------------------
        # IGNORAR FILAS COMPLETAMENTE VACÍAS
        # --------------------------------------------------

        if not any(
            valor is not None and str(valor).strip() != ""
            for valor in fila
        ):

            continue


        total_rows += 1


        try:

            supervisor = fila[col_supervisor]
            tecnico_cedula = fila[col_cedula]
            tecnico_nombre = fila[col_nombre]


            # ==============================================
            # LIMPIAR DATOS
            # ==============================================

            supervisor = (
                str(supervisor).strip()
                if supervisor is not None
                else ""
            )


            # ==============================================
            # LIMPIAR CÉDULA
            # ==============================================

            if tecnico_cedula is None:

                tecnico_cedula = ""

            else:

                if isinstance(
                    tecnico_cedula,
                    float
                ):

                    if tecnico_cedula.is_integer():

                        tecnico_cedula = str(
                            int(tecnico_cedula)
                        )

                    else:

                        tecnico_cedula = str(
                            tecnico_cedula
                        )

                else:

                    tecnico_cedula = str(
                        tecnico_cedula
                    ).strip()


            # ==============================================
            # LIMPIAR NOMBRE
            # ==============================================

            tecnico_nombre = (
                str(tecnico_nombre).strip()
                if tecnico_nombre is not None
                else ""
            )


            # ==============================================
            # VALIDACIONES
            # ==============================================

            errores_fila = {}


            if not supervisor:

                errores_fila["supervisor"] = (
                    "El supervisor es obligatorio."
                )


            if not tecnico_cedula:

                errores_fila["tecnico_cedula"] = (
                    "La cédula del técnico es obligatoria."
                )


            if not tecnico_nombre:

                errores_fila[
                    "tecnico_apellido_nombres"
                ] = (
                    "El nombre del técnico es obligatorio."
                )


            if errores_fila:

                error_count += 1

                errors.append(
                    {
                        "row": numero_fila,
                        "data": {
                            "supervisor": supervisor,
                            "tecnico_cedula": tecnico_cedula,
                            "tecnico_apellido_nombres": (
                                tecnico_nombre
                            ),
                        },
                        "errors": errores_fila,
                    }
                )

                continue


            # ==============================================
            # GUARDAR EN BASE DE DATOS
            # ==============================================

            with transaction.atomic():

                tecnico_existente = (
                    Tecnicos.objects.filter(
                        tecnico_cedula=tecnico_cedula
                    ).first()
                )


                if tecnico_existente:

                    # --------------------------------------
                    # YA EXISTE
                    # --------------------------------------

                    # Si quieres que los repetidos se
                    # ignoren completamente, cambia esta
                    # parte por:
                    #
                    # existing_count += 1
                    # continue

                    cambio = False


                    if (
                        tecnico_existente.supervisor
                        != supervisor
                    ):

                        tecnico_existente.supervisor = (
                            supervisor
                        )

                        cambio = True


                    if (
                        tecnico_existente
                        .tecnico_apellido_nombres
                        != tecnico_nombre
                    ):

                        tecnico_existente.tecnico_apellido_nombres = (
                            tecnico_nombre
                        )

                        cambio = True


                    if cambio:

                        tecnico_existente.save()

                        updated_count += 1

                    else:

                        existing_count += 1


                else:

                    # --------------------------------------
                    # CREAR TÉCNICO
                    # --------------------------------------

                    Tecnicos.objects.create(

                        supervisor=supervisor,

                        tecnico_cedula=tecnico_cedula,

                        tecnico_apellido_nombres=(
                            tecnico_nombre
                        ),

                    )

                    created_count += 1


        except Exception as e:

            error_count += 1

            errors.append(
                {
                    "row": numero_fila,

                    "data": {
                        "supervisor": (
                            str(
                                fila[col_supervisor]
                            )
                            if fila[col_supervisor]
                            is not None
                            else ""
                        ),

                        "tecnico_cedula": (
                            str(
                                fila[col_cedula]
                            )
                            if fila[col_cedula]
                            is not None
                            else ""
                        ),

                        "tecnico_apellido_nombres": (
                            str(
                                fila[col_nombre]
                            )
                            if fila[col_nombre]
                            is not None
                            else ""
                        ),
                    },

                    "errors": {
                        "general": str(e),
                    },
                }
            )


    workbook.close()


    # ======================================================
    # FECHA DE ÚLTIMA CARGA
    # ======================================================

    ahora = timezone.localtime()

    ultima_carga = ahora.strftime(
        "%d/%m/%Y %H:%M:%S"
    )


    # ======================================================
    # GUARDAR RESULTADO EN SESIÓN
    # ======================================================

    request.session["ultima_carga_tecnicos"] = (
        ultima_carga
    )

    request.session["tecnico_report_generated"] = True

    request.session["tecnico_total_rows"] = total_rows

    request.session["tecnico_created_count"] = (
        created_count
    )

    request.session["tecnico_updated_count"] = (
        updated_count
    )

    request.session["tecnico_existing_count"] = (
        existing_count
    )

    request.session["tecnico_error_count"] = (
        error_count
    )

    request.session["tecnico_errors"] = errors


    # ======================================================
    # MENSAJE GENERAL
    # ======================================================

    if error_count == 0:

        messages.success(
            request,
            (
                f"Carga de técnicos completada. "
                f"{created_count} creados, "
                f"{updated_count} actualizados y "
                f"{existing_count} sin cambios."
            ),
        )

    else:

        messages.warning(
            request,
            (
                f"Carga completada con "
                f"{error_count} errores. "
                f"{created_count} técnicos creados."
            ),
        )


    return redirect("carga")

# ==========================================================
# GENERAR PDF DE ERRORES
# ==========================================================

def generar_pdf_errores(request):

    if request.method != "POST":

        return HttpResponse(
            "Método no permitido.",
            status=405
        )

    errores_json = request.POST.get(
        "errores_json"
    )

    if not errores_json:

        return HttpResponse(
            "No se recibieron errores para generar el PDF.",
            status=400
        )

    try:

        error_records = json.loads(
            errores_json
        )

    except json.JSONDecodeError as e:

        return HttpResponse(
            (
                "Los datos de errores no son válidos: "
                f"{e}"
            ),
            status=400
        )

    # ======================================================
    # TEMPLATE
    # ======================================================

    template = get_template(
        "carga/reporte_errores.html"
    )

    html = template.render(
        {
            "errores": error_records,
        }
    )

    # ======================================================
    # RESPUESTA PDF
    # ======================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="reporte_errores_auditorias.pdf"'
    )

    # ======================================================
    # GENERAR PDF
    # ======================================================

    pisa_status = pisa.CreatePDF(
        html,
        dest=response
    )

    if pisa_status.err:

        return HttpResponse(
            "Error al generar el PDF.",
            status=500
        )

    return response


def exportar_excel_completo(request):

    # ==========================================================
    # 1. OBTENER TODA LA BASE DE DATOS
    # ==========================================================

    auditorias = Auditoria.objects.all()

    tecnicos = Tecnicos.objects.all()

    # ==========================================================
    # 2. CREAR LIBRO DE EXCEL
    # ==========================================================

    wb = Workbook()

    # Primera hoja
    ws_estadisticas = wb.active
    ws_estadisticas.title = "Estadísticas"

    # Segunda hoja
    ws_auditorias = wb.create_sheet(
        "BD Auditorías"
    )

    # Tercera hoja
    ws_tecnicos = wb.create_sheet(
        "BD Técnicos"
    )

    # ==========================================================
    # 3. COLORES Y ESTILOS
    # ==========================================================

    COLOR_TITULO = "1F2937"
    COLOR_HEADER = "D9EAF7"
    COLOR_BORDE = "7F8C8D"

    borde = Border(
        left=Side(
            style="thin",
            color=COLOR_BORDE
        ),
        right=Side(
            style="thin",
            color=COLOR_BORDE
        ),
        top=Side(
            style="thin",
            color=COLOR_BORDE
        ),
        bottom=Side(
            style="thin",
            color=COLOR_BORDE
        )
    )

    # ==========================================================
    # 4. AJUSTAR COLUMNAS
    # ==========================================================

    def ajustar_columnas(ws, ancho_maximo=45):

        for columna in range(
            1,
            ws.max_column + 1
        ):

            letra = get_column_letter(
                columna
            )

            maximo = 0

            for celda in ws[letra]:

                if celda.value is not None:

                    longitud = len(
                        str(celda.value)
                    )

                    maximo = max(
                        maximo,
                        longitud
                    )

            ws.column_dimensions[
                letra
            ].width = min(
                maximo + 2,
                ancho_maximo
            )

    # ==========================================================
    # 5. HOJA ESTADÍSTICAS
    # ==========================================================

    ws = ws_estadisticas

    ws.merge_cells("A1:B1")

    ws["A1"] = "ESTADÍSTICAS DE AUDITORÍAS"

    ws["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=14
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=COLOR_TITULO
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    # ==========================================================
    # ENCABEZADOS
    # ==========================================================

    ws["A3"] = "Estadística"
    ws["B3"] = "Cantidad"

    for celda in ws[3]:

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

        celda.alignment = Alignment(
            horizontal="center"
        )

    # ==========================================================
    # ESTADÍSTICAS GENERALES
    # ==========================================================

    total_auditorias = auditorias.count()

    cumplen = auditorias.filter(
        resultado_auditoria="cumple"
    ).count()

    no_cumplen = auditorias.filter(
        resultado_auditoria="no_cumple"
    ).count()

    suspensiones = auditorias.filter(
        tipo_operacion="DC00"
    ).count()

    reconexiones = auditorias.filter(
        tipo_operacion="RC00"
    ).count()

    zvcl = auditorias.filter(
        tipo_operacion="ZVCL"
    ).count()

    tecnicos_count = (
        auditorias
        .exclude(
            nombre_tecnico__isnull=True
        )
        .exclude(
            nombre_tecnico=""
        )
        .values(
            "nombre_tecnico"
        )
        .distinct()
        .count()
    )

    digitadores_count = (
        auditorias
        .exclude(
            nombre_auditor__isnull=True
        )
        .exclude(
            nombre_auditor=""
        )
        .values(
            "nombre_auditor"
        )
        .distinct()
        .count()
    )

    # ==========================================================
    # RESUMEN
    # ==========================================================

    resumen = [

        [
            "Total auditorías",
            total_auditorias
        ],

        [
            "Cumplen",
            cumplen
        ],

        [
            "No cumplen",
            no_cumplen
        ],

        [
            "Suspensiones",
            suspensiones
        ],

        [
            "Reconexiones",
            reconexiones
        ],

        [
            "ZVCL",
            zvcl
        ],

        [
            "Técnicos",
            tecnicos_count
        ],

        [
            "Digitadores",
            digitadores_count
        ],

    ]

    fila = 4

    for nombre, cantidad in resumen:

        ws.cell(
            fila,
            1
        ).value = nombre

        ws.cell(
            fila,
            2
        ).value = cantidad

        ws.cell(
            fila,
            1
        ).border = borde

        ws.cell(
            fila,
            2
        ).border = borde

        ws.cell(
            fila,
            2
        ).alignment = Alignment(
            horizontal="center"
        )

        fila += 1

    # ==========================================================
    # TABLA RESUMEN
    # ==========================================================

    tabla_resumen = Table(
        displayName="TablaResumen",
        ref=f"A3:B{fila - 1}"
    )

    tabla_resumen.tableStyleInfo = (
        TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )

    ws.add_table(
        tabla_resumen
    )

    # ==========================================================
    # 6. AUDITORÍAS POR TÉCNICO
    # ==========================================================

    fila_tecnico = 14

    ws.merge_cells(
        start_row=fila_tecnico,
        start_column=1,
        end_row=fila_tecnico,
        end_column=2
    )

    ws.cell(
        fila_tecnico,
        1
    ).value = "AUDITORÍAS POR TÉCNICO"

    ws.cell(
        fila_tecnico,
        1
    ).font = Font(
        bold=True,
        color="FFFFFF"
    )

    ws.cell(
        fila_tecnico,
        1
    ).fill = PatternFill(
        "solid",
        fgColor=COLOR_TITULO
    )

    ws.cell(
        fila_tecnico,
        1
    ).alignment = Alignment(
        horizontal="center"
    )

    ws.cell(
        fila_tecnico + 1,
        1
    ).value = "Técnico"

    ws.cell(
        fila_tecnico + 1,
        2
    ).value = "Cantidad"

    for columna in range(1, 3):

        celda = ws.cell(
            fila_tecnico + 1,
            columna
        )

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

    estadisticas_tecnicos = (
        auditorias
        .exclude(
            nombre_tecnico__isnull=True
        )
        .exclude(
            nombre_tecnico=""
        )
        .values(
            "nombre_tecnico"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    fila_actual = fila_tecnico + 2

    for tecnico in estadisticas_tecnicos:

        ws.cell(
            fila_actual,
            1
        ).value = tecnico[
            "nombre_tecnico"
        ]

        ws.cell(
            fila_actual,
            2
        ).value = tecnico[
            "total"
        ]

        ws.cell(
            fila_actual,
            1
        ).border = borde

        ws.cell(
            fila_actual,
            2
        ).border = borde

        fila_actual += 1

    # ==========================================================
    # 7. HALLAZGOS
    # ==========================================================

    columna_hallazgo = 4

    ws.merge_cells(
        start_row=fila_tecnico,
        start_column=columna_hallazgo,
        end_row=fila_tecnico,
        end_column=columna_hallazgo + 1
    )

    ws.cell(
        fila_tecnico,
        columna_hallazgo
    ).value = "HALLAZGOS ENCONTRADOS"

    ws.cell(
        fila_tecnico,
        columna_hallazgo
    ).font = Font(
        bold=True,
        color="FFFFFF"
    )

    ws.cell(
        fila_tecnico,
        columna_hallazgo
    ).fill = PatternFill(
        "solid",
        fgColor=COLOR_TITULO
    )

    ws.cell(
        fila_tecnico,
        columna_hallazgo
    ).alignment = Alignment(
        horizontal="center"
    )

    ws.cell(
        fila_tecnico + 1,
        columna_hallazgo
    ).value = "Hallazgo"

    ws.cell(
        fila_tecnico + 1,
        columna_hallazgo + 1
    ).value = "Cantidad"

    for columna in range(
        columna_hallazgo,
        columna_hallazgo + 2
    ):

        celda = ws.cell(
            fila_tecnico + 1,
            columna
        )

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

    hallazgos = (
        auditorias
        .filter(
            resultado_auditoria="no_cumple"
        )
        .exclude(
            hallazgo__isnull=True
        )
        .exclude(
            hallazgo=""
        )
        .values(
            "hallazgo"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    fila_actual_hallazgo = (
        fila_tecnico + 2
    )

    for hallazgo in hallazgos:

        ws.cell(
            fila_actual_hallazgo,
            columna_hallazgo
        ).value = hallazgo[
            "hallazgo"
        ]

        ws.cell(
            fila_actual_hallazgo,
            columna_hallazgo + 1
        ).value = hallazgo[
            "total"
        ]

        ws.cell(
            fila_actual_hallazgo,
            columna_hallazgo
        ).border = borde

        ws.cell(
            fila_actual_hallazgo,
            columna_hallazgo + 1
        ).border = borde

        fila_actual_hallazgo += 1

    # ==========================================================
    # 8. ESTADÍSTICAS POR TIPO DE OPERACIÓN
    # ==========================================================

    columna_operacion = 7

    ws.merge_cells(
        start_row=14,
        start_column=columna_operacion,
        end_row=14,
        end_column=columna_operacion + 1
    )

    ws.cell(
        14,
        columna_operacion
    ).value = "TIPO DE OPERACIÓN"

    ws.cell(
        14,
        columna_operacion
    ).font = Font(
        bold=True,
        color="FFFFFF"
    )

    ws.cell(
        14,
        columna_operacion
    ).fill = PatternFill(
        "solid",
        fgColor=COLOR_TITULO
    )

    ws.cell(
        14,
        columna_operacion
    ).alignment = Alignment(
        horizontal="center"
    )

    ws.cell(
        15,
        columna_operacion
    ).value = "Operación"

    ws.cell(
        15,
        columna_operacion + 1
    ).value = "Cantidad"

    for columna in range(
        columna_operacion,
        columna_operacion + 2
    ):

        celda = ws.cell(
            15,
            columna
        )

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

    operaciones = (
        auditorias
        .values(
            "tipo_operacion"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    fila_operacion = 16

    for operacion in operaciones:

        ws.cell(
            fila_operacion,
            columna_operacion
        ).value = operacion[
            "tipo_operacion"
        ]

        ws.cell(
            fila_operacion,
            columna_operacion + 1
        ).value = operacion[
            "total"
        ]

        ws.cell(
            fila_operacion,
            columna_operacion
        ).border = borde

        ws.cell(
            fila_operacion,
            columna_operacion + 1
        ).border = borde

        fila_operacion += 1

    # ==========================================================
    # 9. ESTADÍSTICAS POR AUDITOR
    # ==========================================================

    columna_auditor = 10

    ws.merge_cells(
        start_row=14,
        start_column=columna_auditor,
        end_row=14,
        end_column=columna_auditor + 1
    )

    ws.cell(
        14,
        columna_auditor
    ).value = "AUDITORÍAS POR AUDITOR"

    ws.cell(
        14,
        columna_auditor
    ).font = Font(
        bold=True,
        color="FFFFFF"
    )

    ws.cell(
        14,
        columna_auditor
    ).fill = PatternFill(
        "solid",
        fgColor=COLOR_TITULO
    )

    ws.cell(
        14,
        columna_auditor
    ).alignment = Alignment(
        horizontal="center"
    )

    ws.cell(
        15,
        columna_auditor
    ).value = "Auditor"

    ws.cell(
        15,
        columna_auditor + 1
    ).value = "Cantidad"

    for columna in range(
        columna_auditor,
        columna_auditor + 2
    ):

        celda = ws.cell(
            15,
            columna
        )

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

    estadisticas_auditores = (
        auditorias
        .exclude(
            nombre_auditor__isnull=True
        )
        .exclude(
            nombre_auditor=""
        )
        .values(
            "nombre_auditor"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    fila_auditor = 16

    for auditor in estadisticas_auditores:

        ws.cell(
            fila_auditor,
            columna_auditor
        ).value = auditor[
            "nombre_auditor"
        ]

        ws.cell(
            fila_auditor,
            columna_auditor + 1
        ).value = auditor[
            "total"
        ]

        ws.cell(
            fila_auditor,
            columna_auditor
        ).border = borde

        ws.cell(
            fila_auditor,
            columna_auditor + 1
        ).border = borde

        fila_auditor += 1

    ajustar_columnas(
        ws,
        45
    )

    ws.freeze_panes = "A3"

    # ==========================================================
    # 10. HOJA BD AUDITORÍAS
    # ==========================================================

    ws = ws_auditorias

    encabezados_auditoria = [

        "ID",
        "Fecha",
        "Nombre Auditor",
        "Número Cédula",
        "Aplicativo",
        "Fecha Operación",
        "Nombre Técnico",
        "Número Cuenta Contrato",
        "Número Orden",
        "Tipo Operación",
        "Resultado Auditoría",
        "Observación",
        "Tipo Hallazgo",
        "Hallazgo",

    ]

    for columna, encabezado in enumerate(
        encabezados_auditoria,
        start=1
    ):

        celda = ws.cell(
            1,
            columna
        )

        celda.value = encabezado

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==========================================================
    # TODAS LAS AUDITORÍAS
    # ==========================================================

    fila = 2

    for auditoria in auditorias:

        datos = [

            auditoria.id,
            auditoria.fecha,
            auditoria.nombre_auditor,
            auditoria.numero_cedula,
            auditoria.aplicativo,
            auditoria.fecha_operacion,
            auditoria.nombre_tecnico,
            auditoria.numero_cuenta_contrato,
            auditoria.numero_orden,
            auditoria.tipo_operacion,
            auditoria.resultado_auditoria,
            auditoria.observacion,
            auditoria.tipo_hallazgo,
            auditoria.hallazgo,

        ]

        for columna, valor in enumerate(
            datos,
            start=1
        ):

            celda = ws.cell(
                fila,
                columna
            )

            celda.value = valor

            celda.border = borde

            celda.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if hasattr(
                valor,
                "strftime"
            ):

                celda.number_format = (
                    "dd/mm/yyyy"
                )

        fila += 1

    # ==========================================================
    # TABLA AUDITORÍAS
    # ==========================================================

    if fila > 2:

        ultima_columna = get_column_letter(
            len(encabezados_auditoria)
        )

        tabla = Table(
            displayName="TablaBDAuditorias",
            ref=(
                f"A1:{ultima_columna}{fila - 1}"
            )
        )

        tabla.tableStyleInfo = (
            TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
        )

        ws.add_table(
            tabla
        )

    ws.freeze_panes = "A2"

    ajustar_columnas(
        ws,
        40
    )

    # ==========================================================
    # 11. HOJA BD TÉCNICOS
    # ==========================================================

    ws = ws_tecnicos

    campos = [
        campo
        for campo in Tecnicos._meta.fields
        if campo.name != "id"
    ]

    encabezados_tecnicos = ["ID"]

    encabezados_tecnicos.extend(
        campo.verbose_name.title()
        for campo in campos
    )

    # ==========================================================
    # ENCABEZADOS
    # ==========================================================

    for columna, encabezado in enumerate(
        encabezados_tecnicos,
        start=1
    ):

        celda = ws.cell(
            1,
            columna
        )

        celda.value = encabezado

        celda.font = Font(
            bold=True
        )

        celda.fill = PatternFill(
            "solid",
            fgColor=COLOR_HEADER
        )

        celda.border = borde

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==========================================================
    # TODOS LOS TÉCNICOS
    # ==========================================================

    fila = 2

    for tecnico in tecnicos:

        valores = [
            tecnico.id
        ]

        valores.extend(
            getattr(
                tecnico,
                campo.name
            )
            for campo in campos
        )

        for columna, valor in enumerate(
            valores,
            start=1
        ):

            celda = ws.cell(
                fila,
                columna
            )

            celda.value = valor

            celda.border = borde

            celda.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if hasattr(
                valor,
                "strftime"
            ):

                celda.number_format = (
                    "dd/mm/yyyy"
                )

        fila += 1

    # ==========================================================
    # TABLA TÉCNICOS
    # ==========================================================

    if fila > 2:

        ultima_columna = get_column_letter(
            len(encabezados_tecnicos)
        )

        tabla = Table(
            displayName="TablaBDTecnicos",
            ref=(
                f"A1:{ultima_columna}{fila - 1}"
            )
        )

        tabla.tableStyleInfo = (
            TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
        )

        ws.add_table(
            tabla
        )

    ws.freeze_panes = "A2"

    ajustar_columnas(
        ws,
        40
    )

    # ==========================================================
    # 12. CONFIGURACIÓN DE IMPRESIÓN
    # ==========================================================

    for hoja in wb.worksheets:

        hoja.page_setup.orientation = (
            "landscape"
        )

        hoja.page_setup.fitToWidth = 1

        hoja.page_setup.fitToHeight = 0

        hoja.sheet_properties.pageSetUpPr.fitToPage = True

    # ==========================================================
    # 13. DESCARGAR EXCEL
    # ==========================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="Reporte_Completo_Auditorias.xlsx"'
    )

    wb.save(response)

    return response